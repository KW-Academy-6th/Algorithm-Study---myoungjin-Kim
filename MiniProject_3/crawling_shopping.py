from selenium import webdriver
import chromedriver_autoinstaller
import time #로딩시간을 기다리기 위해
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By #find_element_by_css_selector 오류를 해결하기위해
from kakaotrans import Translator #번역을 위해
import os # 이미지 저장할 폴더 생성
from urllib.request import urlretrieve # 이미지 다운로드
from openpyxl import Workbook
from openpyxl.drawing.image import Image #엑셀에 이미지를 저장하기 위해

COUNT= 10 #수집할 물품의 개수


#검색할 키워드 입력
query = input('검색할 키워드를 입력하세요: ')

#현재 크롬 드라이버 버전 가져오기
chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]

#현재 크롬 드라이버 버전과 설치된 버전이 다를 경우 맞는 버전으로 업그레이드
try:
    crawler = webdriver.Chrome(f'./{chrome_ver}/chromedriver.exe')
except:
    chromedriver_autoinstaller.install(True)
    crawler = webdriver.Chrome(f'./{chrome_ver}/chromedriver.exe')

crawler.implicitly_wait(10) #time.sleep()과 비슷하지만 창이 생성되면 10초를 기다리지 않고 진행한다.

#해외 사이트를 이용하기 위한 번역 진행
translator = Translator()
query_en = translator.translate(query, src='kr', tgt='en')

#크롬드라이버로 쇼핑몰 url로 접속
url = 'https://ko.aliexpress.com/?spm=a2g0o.productlist.1000002.1.2fc36403fbrcud&gatewayAdapt=glo2kor'
crawler.get(url)
crawler.implicitly_wait(10)

#검색창에 키워드 입력 후 엔터
search_box = crawler.find_element(By.CSS_SELECTOR, "input#search-key") #find_element_by_css_selector에서 오류를 해결하기 위해 변경
search_box.send_keys(query_en)
search_box.send_keys(Keys.ENTER)
crawler.implicitly_wait(10)

#많이 팔린 순 탭 클릭
crawler.find_element(By.XPATH, '//*[@id="root"]/div[1]/div/div[2]/div/div[2]/div[2]/div[1]/div/div/div[2]').click() #동일한 오류 해결을 위해 변경
time.sleep(3) #적정시간 보다 짧을 경우 제대로 로드를 하지 못함

#검색 결과 페이지에서 selenium 패키지로 수집해보기
#find_element_by_css_selector 오류를 해결하기 위해 변경 + 리스트 저장위해서 뒤에 s 추가
product_urls = []
product_names = []
product_prices = []
product_imgs = []

#물품 이름과 가격을 수집합니다.
infos = crawler.find_elements(By.CSS_SELECTOR, ".manhattan--content--1KpBbUi")[:COUNT]
for info in infos:
    try:
        product_names.append(info.find_element(By.CSS_SELECTOR, ".manhattan--titleText--WccSjUS").text)
        product_prices.append(info.find_element(By.CSS_SELECTOR, ".manhattan--price-sale--1CCSZfK").text)
    except:
        print("get name error")

#물품의 url을 수집합니다.
elements = crawler.find_elements(By.XPATH,'//*[@id="root"]/div[1]/div/div[2]/div/div[2]/div[3]/a') #Xpath의 마지막 a뒤의 [1]을 제거하지 않으면 한 개의 값만 반환된다
for element in elements:
    url = element.get_attribute('href')
    product_urls.append(url)


#물품의 이미지를 수집합니다.
imgs = crawler.find_elements(By.CSS_SELECTOR, "img.manhattan--img--36QXbtQ.product-img")[:COUNT]

for img in imgs:    
    product_imgs.append(img.get_attribute('src'))

# 저장할 path_folder의 주소 적기
path_folder = '/Users/kaudw/Pictures/img/'

# 작성된 주소에 폴더가 존재하지 않는다면 생성하기
if not os.path.isdir(path_folder):
    os.mkdir(path_folder)

# 이미지 다운로드
i = 0
for link in product_imgs:          
    i += 1
    urlretrieve(link, path_folder + f'{i}.jpg')

#이미지 파일의 경로를 저장합니다.
image_files = []
for i in range(1, COUNT+1):
    path = path_folder + str(i) + ".jpg"
    image_files.append(path)

# 엑셀 파일 생성
wb = Workbook()

# 시트 생성
ws = wb.active

# 이미지와 리스트를 반복문으로 처리
for i, image_file in enumerate(image_files):
    # 이미지 삽입
    img = Image(image_file)
    ws.add_image(img, f'A{(i+1)*10}')  # 셀 위치 조정

    # 리스트 삽입
    cell_name = ws.cell(row=(i+1)*10, column=5)
    cell_name.value = product_names[i]
    
    cell_price = ws.cell(row=(i+1)*10+1, column=5)
    cell_price.value = product_prices[i]
    
    cell_url = ws.cell(row=(i+1)*10+2, column=5)
    cell_url.value = product_urls[i]
    

# 엑셀 파일 저장
wb.save('./output.xlsx')
